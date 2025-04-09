import os
from collections import defaultdict
from distutils.fancy_getopt import wrap_text

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font
import jsonlines, json
import random
import time
from pathlib import Path
import copy
import torch
import pandas as pd

from auto_evaluators import eval_by_cscore

currtime = time.strftime("%m%d%H%M")


"""
Util: Get path
"""
def get_paths(dir, experiments, primary_seed=2628670643, multi_seed_dict={}):
    image_paths = []
    image_names = []
    topic = dir.split("/")[-1]
    for step in experiments:
        image_names.append(f'{topic}_{step}')
        if step == "original":
            path = f'{dir}/original_{primary_seed}.jpeg'
        elif step == "gpt_rewrite":
            path = f'{dir}/GPT_rewrite_{primary_seed}.jpeg'
        elif step == "dalle3":
            path = f'{dir}/dalle3.jpeg'
        elif step == "gpt_refine":
            path = f'{dir}/gpt_refine/re_0.jpeg'
        elif "abl" in step:
            idx = int(step.split("_")[1])
            path = f'{dir}/abl/re_{idx}.jpeg'
        elif "multi_seed" in step:
            if topic in multi_seed_dict:
                path = multi_seed_dict[topic] 
        elif "seed" in step:
            seed = step.split("_")[-1]
            path = f'{dir}/original_{seed}.jpeg'
        else:
            path = f'{dir}/detailscribe/re_{step}.jpeg'

        if os.path.exists(path):
            image_paths.append(path)
        else:
            print(f'{path} not found.')

    if len(image_paths) != len(experiments):
        image_paths = None
        image_names = None

    return image_paths, image_names


def visualize_by_sheet(jsonl_path, steps, data_list=None, num_rows=25):
    if data_list is None:
        with jsonlines.open(jsonl_path) as reader:
            data_list = [each for each in reader]

    dropdown_options = ["1", "2", "3", "4", "5"]
    dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_options)}"', showDropDown=True)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    ws.title = "Image"

    ws.column_dimensions['A'].width = 40

    for i in range(len(steps)):
        ws.column_dimensions[chr(ord('B') + i * 2 + 1)].width = 10

    # Loop through each image and insert it into the spreadsheet
    for idx in range(len(data_list) // num_rows + 1):
        end_data_idx = len(data_list) if (idx + 1) * num_rows >= len(data_list) else (idx + 1) * num_rows
        for i, each in enumerate(data_list[idx * num_rows: end_data_idx], start=0):
            index = i + 2
            cell = ws[f'A{index}']
            cell.value = each["original_prompt"]
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(size=14)

            for j, image_path in enumerate(each["image_paths"]):
                img = Image(image_path)
                img.width, img.height = 256, 256  # Resize image for better fit (optional)
                # Insert image name and image in the spreadsheet

                ws.add_image(img, f"{chr(ord('B') + j * 2)}{index}")
                ws.column_dimensions[chr(ord('B') + j * 2)].width = img.width * 0.13
                ws.row_dimensions[index].height = img.height * 0.75
                # Apply dropdown to the "C" column
                cell_with_dropdown = f"{chr(ord('B') + j * 2 + 1)}{index}"
                dv.add(ws[cell_with_dropdown])
        ws.add_data_validation(dv)
        # Save the workbook
        output_file = jsonl_path.replace(".jsonl", f"_{idx + 1:02d}.xlsx")
        wb.save(output_file)
        print(f'Spreadsheet saved as {output_file}')
    return


def prompt_to_jsonl(jsonl_path, category, multi_seed_path=None, seed=2628670643, type="main",
                   steps=[0, 1, 2, 4], sheet=False):
    experiments = []
    data_list = []
    if type == "main":
        out_path = f'results/eval/dict/main_{category}.jsonl'
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        experiments = list(map(str, steps)) + ["original", "gpt_rewrite", "dalle3", "gpt_refine", "multi_seed"]
        with jsonlines.open(jsonl_path) as reader:
            num_of_prompt = 0
            multi_seed = {}
            if multi_seed_path:
                with jsonlines.open(multi_seed_path) as multi_seed_reader:
                    multi_seed = {each["topic"]: each["best_path"] for each in multi_seed_reader}
            for each in reader:
                # shuffle the image order to reduce the bias of human evaluation
                shuffled_experiments = random.sample(experiments, len(experiments))
                image_paths, _ = get_paths(each["dir"], shuffled_experiments, seed, multi_seed_dict=multi_seed)
                if image_paths:
                    each["image_paths"] = image_paths
                    each["steps"] = shuffled_experiments
                    data_list.append(each)
                    num_of_prompt += 1
            print(f'Read in {num_of_prompt} prompts from {jsonl_path}.')

    elif type == "ab_decompose":
        out_path = f'results/eval/dict/abl_decompose_{category}.jsonl'
        Path(out_path.parent).mkdir(parents=True, exist_ok=True)
        for step in steps:
            experiments.append(str(step))
            experiments.append(f'abl_{str(step)}')
        with jsonlines.open(jsonl_path) as reader:
            for each in reader:
                shuffled_experiments = random.sample(experiments, len(experiments))
                image_paths = get_paths(each["dir"], shuffled_experiments, seed)
                # for sample in experiments:
                #     if "abl" in sample:
                #         idx = int(sample.split("_")[1])
                #         image_paths.append(f'{dir}/abl/re_{idx}.jpeg')
                #     else:
                #         image_paths.append(f'{dir}/detailscribe/re_{sample}.jpeg')
                if image_paths:
                    each["image_paths"] = image_paths
                    each["steps"] = shuffled_experiments
                    data_list.append(each)

    # write data_list to jsonl for log
    with jsonlines.open(out_path, "w") as writer:
        for each in data_list:
            writer.write(each)
    if sheet:
        visualize_by_sheet(out_path, data_list, experiments, num_rows=25)

    return

def rating_per_topic(image_path_dict, all_scores, topics, seeds):
    groups = [f"seed_{seed}" for seed in seeds]
    best_image_path_per_topic = {}
    for i, topic in enumerate(topics):
        scores = []
        local_id = []
        for j, group in enumerate(groups):
            image_id = f'{topic}_{group}'
            local_id.append(image_id)
            scores.append(all_scores[image_id])
        # find the index of max value in list
        max_index = scores.index(max(scores))
        best_image_path_per_topic[topic] = image_path_dict[local_id[max_index]]
    return best_image_path_per_topic

def multi_seed(eval_jsonl, seeds, sheet=False):
    assert len(seeds) > 1
    image_path_dict = {}
    image_ids = []
    image_paths = []
    cleaned_data_list = []
    captions = {}
    topics = {}
    experiments = [f"seed_{seed}" for seed in seeds]
    with jsonlines.open(eval_jsonl) as reader:
        for each in reader:
            paths, image_names = get_paths(each["dir"], experiments)
            print(paths)
            if paths:
                # filter out bad topic (missing images)
                each["steps"] = experiments
                cleaned_data_list.append(each)
                topics[each["topic"]] = 1
                for path, image_name in zip(paths, image_names):
                    image_ids.append(image_name)
                    image_paths.append(path)
                    image_path_dict[image_name] = path
                    captions[image_name] = each['original_prompt']
    topics = topics.keys()
    candidates = captions
    candidates = [candidates[cid] for cid in image_ids]
    device = "cuda" if torch.cuda.is_available() else "cpu"
    scores = eval_by_cscore(image_ids, image_paths, candidates, device)

    best_image_path_per_topic = rating_per_topic(image_path_dict, scores, topics, seeds)

    out_path = f'{Path(eval_jsonl).parent}/multi_seed_' + '_'.join(str(seed) for seed in seeds) + '.jsonl'
    print(f'writing to {out_path}')
    with jsonlines.open(out_path, "w") as writer:
        for each in cleaned_data_list:
            each["best_path"] = best_image_path_per_topic[each["topic"]]
            writer.write(each)

    if sheet:
        visualize_by_sheet(out_path, cleaned_data_list, experiments, num_rows=25)

    return out_path

def stats_one_sheet(sheet_dir, sheet_name, log_path, type="main",
              annotator_list=[], stats_dir='results/eval/stats'):
    # Load the workbook and the specific sheet
    wb_list = [load_workbook(f'{sheet_dir}/{annotator}/{sheet_name}.xlsx', data_only=True) for annotator in
               annotator_list]
    ws_list = [wb["Sheet1"] for wb in wb_list]

    if type == "main":
        groups = ["original", "gpt_rewrite", "dalle3", "multi_seed", "gpt_refine", "0", "1", "2", "4"]
        methods = ["SD", "+GPT Rewrite", "DALLE3", "+multi-seed", "+GPT Refine", "0", "1", "2", "4",
                   "DetailScribe"]
    elif type == "ab_decompose":
        groups = ["abl_0", "abl_1", "abl_2", "abl_4", "0", "1", "2", "4"]
        methods = ["abl", "DetailScribe"]
    else:
        assert 0
    sheets_list = [defaultdict(list) for _ in range(len(annotator_list))]

    all_ratings_df = []
    ratings_by_prompt = {}
    """
    {"prompt0": {"Annotator0": {"SD": 1, "+GPT Rewrite": 2, "DALLE3": 3, "+multi-seed": 4, "+GPT Refine": 5, "0": 6, "1": 7, "2": 8, "4": 9, "DetailScribe": 10},
                "Annotator1": {"SD": 1, "+GPT Rewrite": 2, "DALLE3": 3, "+multi-seed": 4, "+GPT Refine": 5, "0": 6, "1": 7, "2": 8, "4": 9, "DetailScribe": 10},
                ...}, "prompt1"
    """

    with jsonlines.open(log_path) as reader:
        total_game = 0
        for i, each in enumerate(reader):
            # if each["original_prompt"] in ["A tortoise riding a bike made from leaves and twigs, slowly making its way through a forest path.",
            #                                "An owl sitting at a desk, holding a quill in its beak, writing in an open book under the light of a small candle.", # manipulation
            #                                "A fluffy cloud made from cotton candy with small blue gummy form raindrops falling from it.",
            #                                "A solar system made of colored candies as planets, with a glowing candle as the sun.",
            #                                "A clock face using different colored fruits to mark different hours.", # abstract concept
            #                                "Two elephants use their trunks to carry a large log together, walking in sync to balance the weight.",
            #                                "Two squirrels climb a tree trunk, helping each other up by nudging and stabilizing one another.",
            #                                "A heron and a bear fish side-by-side in a stream, each catching their dinner.",
            #                                "A beaver and a turtle work together to lift a heavy stone to build a dam."]:
            #     continue
            total_game += 1
            index = i + 2
            shuffled_steps = each["steps"]
            # get an ordered score
            # get a total win rate
            votes = []
            for k, ws in enumerate(ws_list):
                scores = []
                for j, group in enumerate(groups):
                    col = shuffled_steps.index(group)
                    cell = ws[f'{chr(ord("B") + col * 2 + 1)}{index}']
                    score = cell.value
                    scores.append(score)
                    sheets_list[k][f'{group}'].append(score)
                if type == "main":
                    new_scores = [*scores, max(scores[-4:])]
                elif type == "ab_decompose":
                    new_scores = [max(scores[:4]), max(scores[4:])]
                for j, ss in enumerate(new_scores):
                    all_ratings_df.append({
                        "example_id": i,
                        "prompt": each["original_prompt"],
                        "annotator_id": k,
                        "method": methods[j],
                        "score": ss,
                        "is_win": ss == np.max(new_scores)
                    })
                    if each["original_prompt"] not in ratings_by_prompt:
                        ratings_by_prompt[each["original_prompt"]] = {}
                    if f'annotator_{k}' not in ratings_by_prompt[each["original_prompt"]]:
                        ratings_by_prompt[each["original_prompt"]][f'annotator_{k}'] = {}
                    ratings_by_prompt[each["original_prompt"]][f'annotator_{k}'][methods[j]] = ss
                    print(f'add {methods[j]} to ratings_by_prompt.')
                ratings_by_prompt[each["original_prompt"]][f'annotator_{k}'] = dict(sorted(ratings_by_prompt[each["original_prompt"]][f'annotator_{k}'].items()))
    score_dir = f'{stats_dir}/{sheet_name}_human'
    Path(score_dir).mkdir(parents=True, exist_ok=True)

    with open(f'{score_dir}/human_score_by_prompt.json', 'w') as f:
        json.dump(ratings_by_prompt, f, indent=4)
    all_ratings_df = pd.DataFrame(all_ratings_df)
    print(f'Total ratings: {len(all_ratings_df)}')
    print("=== Rater Scores Mean Std")
    print(all_ratings_df.groupby("annotator_id").aggregate({
        # "annotator_name": "first",
        "score": ["mean", "std", "sem"]
    }))

    print("\n\n=== Method Scores Mean Std")
    print(all_ratings_df.groupby("method").aggregate({
        "score": ["mean", "std", "sem"],
    }))

    print("\n\n=== Win Rate Mean Std")
    print(all_ratings_df.groupby(["annotator_id", "method"]).aggregate({
        "is_win": "mean"
    }).reset_index().groupby("method").aggregate({
        "is_win": ["mean", "std", "sem"]
    }))

    # by_prompt_df = all_ratings_df.groupby(["example_id", "method"], as_index=False).agg(
    #     {"score": "mean"})
    by_prompt_df = all_ratings_df.groupby(["prompt", "method"], as_index=False).agg(
        {"score": "mean"})

    stats_path = f'{sheet_dir}/stats_{sheet_name}.json'
    print(f'writing to {stats_path}')
    all_ratings_df.to_json(stats_path, orient="records", lines=True)

    stats_by_prompt_path = f'{score_dir}/avg_human.json'
    print(f'writing to {stats_by_prompt_path}')
    by_prompt_df.to_json(stats_by_prompt_path, orient="records", lines=True)

    print("\n\n=== prompt Scores Mean Std")
    print(all_ratings_df.groupby(["example_id", "method"])["score"].mean().reset_index())
    with open(f'{score_dir}/human_score_by_prompt.txt', 'w') as f:
        # f.write(all_ratings_df.groupby(["example_id", "method"])["score"].mean().reset_index().to_string())
        f.write(all_ratings_df.groupby(["prompt", "method"])["score"].mean().reset_index().to_string())

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--new_prompt_path", type=str,
                        default=None)
    parser.add_argument("--data_name", type=str, default=None)  # category
    parser.add_argument("--primary_seed", default=2628670643, type=int)
    parser.add_argument("--multi_seed", default=None, type=int, nargs='+')
    parser.add_argument("--sheet", action="store_true")
    parser.add_argument("--ablation", action="store_true")  # For ablation study

    # For manual rating reads in
    parser.add_argument("--read_ratings", action="store_true")  # read in ratings from excel
    parser.add_argument("--rating_dir", type=str, default=None)
    parser.add_argument("--sheet_name", type=str, default=None)
    parser.add_argument("--log_path", type=str, default=None)
    parser.add_argument("--annotator_list", type=str, default=None, nargs='+')  # list of annotators to read in

    args = parser.parse_args()
    eval_type = "main" if not args.ablation else "ab_decompose"
    if args.read_ratings:
        assert args.rating_dir is not None
        assert args.sheet_name is not None
        assert args.log_path is not None
        assert args.annotator_list is not None
        stats_one_sheet(args.rating_dir, args.sheet_name, args.log_path, type=eval_type, annotator_list=args.annotator_list)
        exit(0)

    if args.data_name is None:
        args.data_name = args.new_prompt_path.split("/")[-2]
    multi_seed_path = multi_seed(args.new_prompt_path, [args.primary_seed, *args.multi_seed], sheet=args.sheet) if args.multi_seed else None
    prompt_to_jsonl(args.new_prompt_path, category=args.data_name, multi_seed_path=multi_seed_path, type=eval_type, sheet=args.sheet)
